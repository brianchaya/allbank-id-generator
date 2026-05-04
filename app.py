import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import PatternFill

st.title("All Bank ID Generator")

mode = st.radio(
    "Pilih jenis file",
    ["1 File (RK + Database)", "2 File (RK dan Database terpisah)"]
)

rk_file = None
db_file = None

if mode == "1 File (RK + Database)":
    rk_file = st.file_uploader("Upload file RK + Database", type=["xlsx"])
else:
    rk_file = st.file_uploader("Upload Rekening Koran", type=["xlsx"])
    db_file = st.file_uploader("Upload Database", type=["xlsx"])

# =====================================
# DETECT HEADER (FIXED)
# =====================================
def detect_header(df):

    best_row = 0
    max_valid = 0

    for i in range(min(20, len(df))):

        row = df.iloc[i]

        # hitung berapa banyak cell yang "kayak header"
        valid_count = sum([
            isinstance(x, str) and len(x.strip()) > 0
            for x in row
        ])

        # pilih baris dengan isi terbanyak (biasanya header)
        if valid_count > max_valid:
            max_valid = valid_count
            best_row = i

    return best_row


# =====================================
# DETECT RK SHEET
# =====================================
def detect_rk_sheet(excel):

    for sheet in excel.sheet_names:

        # Pakai header=None supaya bisa scan semua cell, bukan cuma nama kolom
        df = pd.read_excel(excel, sheet_name=sheet, header=None, nrows=20)

        # Cek semua cell di 20 baris pertama
        all_text = df.astype(str).values.flatten()

        if any(
            "uraian" in str(x).lower() or
            "description" in str(x).lower() or
            "keterangan" in str(x).lower()
            for x in all_text
        ):
            return sheet

    return excel.sheet_names[0]

# =====================================
# DETECT DATABASE SHEET
# =====================================
def detect_db_sheet(excel):

    for sheet in excel.sheet_names:

        df = pd.read_excel(excel, sheet_name=sheet, nrows=10)

        cols = [str(c).lower() for c in df.columns]

        if "id" in cols:
            return sheet

    return excel.sheet_names[-1]


# =====================================
# DETECT TRANSACTION COLUMN
# =====================================
def detect_transaction_col(df):

    for col in df.columns:

        name = str(col).lower()

        if "uraian" in name \
        or "description" in name \
        or "keterangan" in name:

            return col

    lengths = df.astype(str).apply(lambda x: x.str.len().mean())
    return lengths.idxmax()


# =====================================
# DETECT DB COLUMN
# =====================================
def detect_db_columns(db):

    kode_col = None
    id_col = None

    for col in db.columns:

        name = str(col).lower()

        if "kode" in name:
            kode_col = col

        if name == "id":
            id_col = col

    return kode_col, id_col


# =====================================
# FAST SEARCH
# =====================================
def generate_ids(text_series, kode_list, id_list):

    pairs = []
    for kode, id_val in zip(kode_list, id_list):
        sub_kodes = [k.strip() for k in str(kode).split(";")]
        for sk in sub_kodes:
            if sk not in ("", "nan", "none", "N/A", "n/a"):
                pairs.append((sk, id_val))

    pairs.sort(key=lambda x: len(str(x[0])), reverse=True)

    results = []
    is_double_id = []

    for text in text_series:
        if pd.isna(text):
            results.append(None)
            is_double_id.append(False)
            continue

        found_ids = []
        text_upper = ' ' + str(text).upper() + ' '

        for kode, id_val in pairs:
            try:
                kode_upper = ' ' + str(kode).upper().strip() + ' '
                if kode_upper in text_upper:
                    id_str = str(id_val)
                    if id_str not in found_ids:
                        found_ids.append(id_str)
            except Exception:
                continue

        final_id = " ; ".join(found_ids) if found_ids else None
        results.append(final_id)
        is_double_id.append(len(found_ids) > 1)

    return results, is_double_id

# =====================================
# MAIN PROCESS
# =====================================
if mode == "1 File (RK + Database)" and rk_file:

    excel = pd.ExcelFile(rk_file)

    rk_sheet = detect_rk_sheet(excel)
    db_sheet = detect_db_sheet(excel)

    preview = pd.read_excel(excel, sheet_name=rk_sheet, header=None)
    header_row = detect_header(preview)

    rk = pd.read_excel(excel, sheet_name=rk_sheet, header=header_row)
    
    # FIX HEADER DOUBLE / UNNAMED ROW
    # =====================================
    
    # kalau baris pertama isinya mirip nama kolom → berarti itu header nyasar
    first_row = rk.iloc[0].astype(str).str.lower().tolist()
    col_names = [str(c).lower() for c in rk.columns]
    
    match_count = sum([1 for x in first_row if x in col_names])
    
    # kalau banyak yg sama → buang baris itu
    if match_count >= len(col_names) // 2:
        rk = rk.iloc[1:].reset_index(drop=True)
        
    db = pd.read_excel(excel, sheet_name=db_sheet)

    wb = load_workbook(rk_file)
    ws = wb[rk_sheet]


elif mode == "2 File (RK dan Database terpisah)" and rk_file and db_file:

    excel_rk = pd.ExcelFile(rk_file)
    rk_sheet = detect_rk_sheet(excel_rk)

    preview = pd.read_excel(excel_rk, sheet_name=rk_sheet, header=None)
    header_row = detect_header(preview)

    rk = pd.read_excel(excel_rk, sheet_name=rk_sheet, header=header_row)

    excel_db = pd.ExcelFile(db_file)
    db_sheet = detect_db_sheet(excel_db)

    db = pd.read_excel(excel_db, sheet_name=db_sheet)

    wb = load_workbook(rk_file)
    ws = wb[rk_sheet]

else:
    st.stop()


# =====================================
# DETECT COLUMN
# =====================================
desc_col = detect_transaction_col(rk)
kode_col, id_col = detect_db_columns(db)

if kode_col is None or id_col is None:
    st.error("Unique code or ID column not found in database")
    st.stop()

kode_list = db[kode_col].astype(str).tolist()
id_list = db[id_col].tolist()


# =====================================
# GENERATE IDS
# =====================================
rk["ID"], is_double_id = generate_ids(rk[desc_col], kode_list, id_list)

st.subheader("Preview Hasil")
st.dataframe(rk)

# =====================================
# WRITE BACK TO EXCEL
# =====================================
header_row_excel = header_row + 1

while ws.cell(header_row_excel, 1).value is None:
    header_row_excel += 1

id_col_excel = None

for c in range(1, ws.max_column + 1):
    val = ws.cell(header_row_excel, c).value
    if val and str(val).strip().lower() == "id":
        id_col_excel = c
        break

if id_col_excel is None:
    id_col_excel = ws.max_column + 1
    ws.cell(header_row_excel, id_col_excel).value = "ID"

red_fill  = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
blue_fill = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")

for i, (val, double_flag) in enumerate(zip(rk["ID"], is_double_id)):
    excel_row = header_row_excel + 1 + i
    cell = ws.cell(excel_row, id_col_excel)
    cell.value = val

    if pd.isna(val) or val is None:
        cell.fill = red_fill
    elif double_flag:
        cell.fill = blue_fill


# =====================================
# DOWNLOAD
# =====================================
output = BytesIO()
wb.save(output)

# =====================================
# SUMMARY INFO
# =====================================
blank_count = rk["ID"].isna().sum()
total_rows = len(rk)

st.warning(f"{blank_count} IDs were not found (out of {total_rows} data)")

st.download_button(
    "Download RK with ID",
    output.getvalue(),
    "RK_HASIL_ID.xlsx"
)
